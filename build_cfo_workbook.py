"""
build_cfo_workbook.py
─────────────────────────────────────────────────────────────────────────────
PACES Episode-of-Care CFO Workbook Generator
Colorado Medicaid · KPMG Engagement · Episodes of Care Solutions

Takes a Colorado APCD facility-episode risk-adjusted cost CSV as input and
produces a 10-worksheet Excel workbook organized for CFO/HCPF decision-making.

USAGE
─────────────────────────────────────────────────────────────────────────────
  python build_cfo_workbook.py INPUT.csv
  python build_cfo_workbook.py INPUT.csv --output OUTPUT.xlsx
  python build_cfo_workbook.py INPUT.csv --config config.json
  python build_cfo_workbook.py INPUT.csv --min-volume 5 --oe-threshold 1.10
  python build_cfo_workbook.py --help

OUTPUT WORKSHEETS
─────────────────────────────────────────────────────────────────────────────
  1.  README                 — methodology, definitions, key totals
  2.  All Data               — complete dataset with computed analytical columns
  3.  Analytical View        — filtered to volume >= threshold
  4.  Outliers Only          — O:E > threshold within analytical view
  5.  Systemic Outliers      — facilities outlier in 3+ episodes
  6.  Top-30 Concentration   — facility-episodes ranked by outlier excess $
  7.  Episode Summary        — per-episode rollup with classification
  8.  By RAE Region          — outlier excess concentration by RAE catchment
  9.  By DOI Urbanity        — outlier rate and excess by facility category
 10.  By Hospital System     — outlier excess by system affiliation

DEPENDENCIES
─────────────────────────────────────────────────────────────────────────────
  Python 3.9+
  openpyxl >= 3.0   (pip install openpyxl)
  No other dependencies. Standard library for CSV parsing.

INPUT REQUIREMENTS
─────────────────────────────────────────────────────────────────────────────
  See INPUT_SPECIFICATION.md in the same folder. Briefly:

  Required columns (case-sensitive header names):
    Episode Name, Facility Name, Facility Hospital System,
    Facility RAE Region, Facility DOI Category, Medicaid Volume,
    High Risk Volume, Total Actual Cost, Facility O:E Cost Ratio

  Optional but used if present:
    Clinical Chapter, Facility County, Facility HSR,
    CDPS Facility Risk Score, all base-rate columns

CONFIGURATION
─────────────────────────────────────────────────────────────────────────────
  Episode classifications (Standard / Heterogeneous) and addressable %
  defaults are externalized to config.json (same folder). Edit that file
  to add new episodes or adjust addressable assumptions; no code change
  required. If an episode appears in the input that is not in config.json,
  the script defaults to Heterogeneous and logs a warning.
─────────────────────────────────────────────────────────────────────────────
"""

import argparse
import csv
import json
import sys
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ────────────────────────────────────────────────────────────────────────────
# DEFAULT CONFIGURATION (overrideable via --config config.json)
# ────────────────────────────────────────────────────────────────────────────
DEFAULT_CONFIG = {
    "episode_classifications": {
        "CABG (including Cardiac catheterization)": "Standard",
        "Coronary Artery Bypass Graft (CABG)": "Standard",
        "Open heart valve surgery (including Cardiac catheterization)": "Standard",
        "Percutaneous cardiac intervention (including Cardiac catheterization)": "Standard",
        "Thyroidectomy": "Standard",
        "Esophagogastroduodenoscopy (Upper Endoscopy)": "Heterogeneous",
        "Cholecystectomy": "Heterogeneous",
        "Bariatric Surgery": "Heterogeneous",
        "Mastectomy": "Heterogeneous",
        "Colectomy": "Heterogeneous",
        "Repair Ventral Hernia": "Heterogeneous",
        "Leg Amputation": "Heterogeneous",
    },
    "addressable_pct_by_classification": {
        "Standard": 0.70,
        "Heterogeneous": 0.50,
    },
    "default_classification_for_unknown": "Heterogeneous",
    "rae_descriptors": {
        "1": "Northeast / Western Slope",
        "2": "Boulder / Larimer",
        "3": "Pueblo / Colorado Springs",
        "4": "Denver Metro / Adams / Arapahoe",
    },
    "min_volume_filter": 5,
    "oe_outlier_threshold": 1.10,
    "systemic_min_episodes": 3,
    "top_n_concentration": 30,
}

# Required column names (case-sensitive)
REQUIRED_COLUMNS = [
    "Episode Name",
    "Facility Name",
    "Facility Hospital System",
    "Facility RAE Region",
    "Facility DOI Category",
    "Medicaid Volume",
    "High Risk Volume",
    "Total Actual Cost",
    "Facility O:E Cost Ratio",
]

# Styling constants
NAVY = "1A3A5C"
NAVY_DK = "0F2540"
GOLD = "E8A000"
LIGHT_BG = "F3F7FB"
RED_BG = "FCE8E6"
GREEN_BG = "E8F5E9"
YELLOW_BG = "FFFDE7"


# ────────────────────────────────────────────────────────────────────────────
# UTILITIES
# ────────────────────────────────────────────────────────────────────────────
def parse_number(s):
    """Parse a string that may contain $ , % or whitespace into a float."""
    if s is None:
        return 0.0
    s = str(s).strip().replace(",", "").replace("$", "").replace("%", "")
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def validate_schema(header_row, required_columns):
    """Check that the input file has the required columns. Returns column index map."""
    col_idx = {}
    missing = []
    for col in required_columns:
        if col in header_row:
            col_idx[col] = header_row.index(col)
        else:
            missing.append(col)
    if missing:
        print("ERROR: Input file is missing required column(s):", file=sys.stderr)
        for col in missing:
            print(f"  - {col!r}", file=sys.stderr)
        print(
            "\nExpected exact case-sensitive header names. See INPUT_SPECIFICATION.md.",
            file=sys.stderr,
        )
        sys.exit(2)
    return col_idx


def load_config(config_path):
    """Load JSON config, layering over DEFAULT_CONFIG."""
    config = dict(DEFAULT_CONFIG)
    if config_path and Path(config_path).exists():
        with open(config_path, encoding="utf-8") as f:
            user_config = json.load(f)
        for k, v in user_config.items():
            if isinstance(v, dict) and isinstance(config.get(k), dict):
                config[k] = {**config[k], **v}
            else:
                config[k] = v
        print(f"  Loaded config from: {config_path}")
    return config


def read_input_csv(path, col_idx):
    """Read the input CSV and return list of dict-like records."""
    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)
    header = rows[0]
    # Filter to non-blank rows (Episode Name populated)
    ep_idx = col_idx["Episode Name"]
    data = [r for r in rows[1:] if len(r) > ep_idx and r[ep_idx].strip()]
    return header, data


# ────────────────────────────────────────────────────────────────────────────
# CORE ANALYTICS
# ────────────────────────────────────────────────────────────────────────────
def compute_oe_ranks_within_episode(data, col_idx):
    """Rank facilities within each episode by O:E ratio ascending (1 = best)."""
    by_ep = defaultdict(list)
    oe_idx = col_idx["Facility O:E Cost Ratio"]
    ep_idx = col_idx["Episode Name"]
    for i, r in enumerate(data):
        by_ep[r[ep_idx].strip()].append((i, parse_number(r[oe_idx])))
    ranks = {}
    for ep, items in by_ep.items():
        items_sorted = sorted(items, key=lambda x: x[1])
        for rank, (orig_i, _) in enumerate(items_sorted, start=1):
            ranks[orig_i] = rank
    return ranks


def augment_row(orig_idx, row, col_idx, config, oe_ranks):
    """Add the six computed analytical columns to a row."""
    vol = parse_number(row[col_idx["Medicaid Volume"]])
    hr = parse_number(row[col_idx["High Risk Volume"]])
    oe = parse_number(row[col_idx["Facility O:E Cost Ratio"]])
    actual = parse_number(row[col_idx["Total Actual Cost"]])
    ep_name = row[col_idx["Episode Name"]].strip()

    pct_hr = (hr / vol * 100) if vol > 0 else 0
    cls = config["episode_classifications"].get(
        ep_name, config["default_classification_for_unknown"]
    )

    outlier_threshold = config["oe_outlier_threshold"]
    is_outlier = oe > outlier_threshold
    excess = (actual * (1 - 1 / oe)) if (oe > 0 and is_outlier) else 0

    flags = []
    if vol == 0:
        flags.append("zero_vol")
    elif vol < 3:
        flags.append("very_low_vol")
    elif vol < config["min_volume_filter"]:
        flags.append("low_vol")
    if oe == 0:
        flags.append("no_OE")
    dq = ";".join(flags) if flags else "OK"

    return list(row) + [
        cls,
        f"{pct_hr:.1f}%",
        oe_ranks.get(orig_idx, 0),
        f"{excess:.2f}" if is_outlier else "",
        "Yes" if is_outlier else "No",
        dq,
    ]


def warn_unclassified_episodes(data, col_idx, config):
    """Warn if input contains episodes not in the classification config."""
    classified = set(config["episode_classifications"].keys())
    found = {r[col_idx["Episode Name"]].strip() for r in data}
    unknown = found - classified
    if unknown:
        default = config["default_classification_for_unknown"]
        print("\n  WARNING: Episodes not in classification config:")
        for ep in sorted(unknown):
            print(f"    - {ep!r}  -> defaulted to {default!r}")
        print(
            "  Update config.json -> 'episode_classifications' "
            "to classify these episodes explicitly.\n"
        )


# ────────────────────────────────────────────────────────────────────────────
# STYLE HELPERS
# ────────────────────────────────────────────────────────────────────────────
def thin_border():
    side = Side(style="thin", color="C8D8EA")
    return Border(left=side, right=side, top=side, bottom=side)


def header_fill():
    return PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")


def style_header_row(ws, row_num, ncols):
    bord = thin_border()
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bord


def autosize_columns(ws, max_width=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                v = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(v))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)


def filter_banner(ws, text, color_bg, color_fg, ncols):
    ws["A1"] = text
    ws["A1"].font = Font(bold=True, italic=True, size=10, color=color_fg)
    ws["A1"].fill = PatternFill(
        start_color=color_bg, end_color=color_bg, fill_type="solid"
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)


# ────────────────────────────────────────────────────────────────────────────
# WORKSHEET BUILDERS
# ────────────────────────────────────────────────────────────────────────────
def build_readme(wb, totals, config):
    ws = wb.active
    ws.title = "README"

    content = [
        ("EPISODE-OF-CARE COST RISK-ADJUSTED ANALYSIS WORKBOOK", "title"),
        ("Colorado Medicaid · APCD-derived facility-episode performance", "subtitle"),
        ("Generated by build_cfo_workbook.py · Episodes of Care Solutions", "meta"),
        ("", ""),
        ("WORKSHEET INDEX", "section"),
        ("1. README — this sheet, methodology and definitions", "body"),
        ("2. All Data — complete dataset with added analytical columns", "body"),
        (
            f"3. Analytical View — filtered to Medicaid volume >= "
            f"{config['min_volume_filter']}",
            "body",
        ),
        (
            f"4. Outliers Only — O:E > {config['oe_outlier_threshold']} within "
            "analytical view",
            "body",
        ),
        (
            f"5. Systemic Outliers — facilities outlier in "
            f"{config['systemic_min_episodes']}+ episodes",
            "body",
        ),
        (
            f"6. Top-{config['top_n_concentration']} Concentration — "
            "facility-episodes ranked by outlier excess $",
            "body",
        ),
        ("7. Episode Summary — per-episode rollup with classification & addressable", "body"),
        ("8. By RAE Region — outlier excess concentration by RAE catchment", "body"),
        ("9. By DOI Urbanity — outlier rate and excess by facility category", "body"),
        ("10. By Hospital System — outlier excess by system affiliation", "body"),
        ("", ""),
        ("FILTERS APPLIED ACROSS WORKSHEETS", "section"),
        (
            f"- Volume filter: Medicaid Volume >= {config['min_volume_filter']} "
            "(suppresses small-n statistical noise)",
            "body",
        ),
        (
            f"- Outlier threshold: Facility O:E Cost Ratio > "
            f"{config['oe_outlier_threshold']} (>{int((config['oe_outlier_threshold']-1)*100)}% above expected)",
            "body",
        ),
        (
            "- Both filters applied to Worksheets 3 through 10. Worksheet 2 retains all rows.",
            "body",
        ),
        ("", ""),
        ("ADDED ANALYTICAL COLUMNS (on Worksheets 2-4)", "section"),
        ("- Episode Classification — Standard or Heterogeneous", "body"),
        (
            "    Standard: tight procedure list, diagnosis not the driver "
            "(CABG, Valves, PCI, Thyroidectomy)",
            "body",
        ),
        (
            "    Heterogeneous: bundled procedures or diagnoses with case-mix variance "
            "(EGD, Mastectomy, Colectomy, etc.)",
            "body",
        ),
        (
            "- % High Risk Volume — High Risk Volume / Medicaid Volume "
            "(acuity proxy)",
            "body",
        ),
        (
            "- O:E Rank (within Episode) — rank 1 = lowest O:E (best); "
            "rank N = highest O:E (worst)",
            "body",
        ),
        (
            "- Outlier Excess $ — for outliers only, computed as "
            "Actual * (1 - 1/O:E)",
            "body",
        ),
        ("- Is Outlier (O:E>1.10) — Yes / No flag", "body"),
        (
            "- Data Quality Flag — low_vol, very_low_vol, zero_vol, no_OE, or OK",
            "body",
        ),
        ("", ""),
        ("ADDRESSABLE EXCESS METHODOLOGY", "section"),
        (
            "Not all outlier excess is recoverable through facility QI engagement.",
            "body",
        ),
        (
            "Heterogeneous episodes lump together different procedures or diagnoses,",
            "body",
        ),
        (
            "so part of the cost variance reflects definitional/case-mix issues that",
            "body",
        ),
        ("engagement cannot resolve. The defaults applied:", "body"),
        (
            f"    Standard episodes: "
            f"{int(config['addressable_pct_by_classification']['Standard']*100)}% "
            "of outlier excess is addressable",
            "body",
        ),
        (
            f"    Heterogeneous episodes: "
            f"{int(config['addressable_pct_by_classification']['Heterogeneous']*100)}% "
            "of outlier excess is addressable",
            "body",
        ),
        (
            "Both percentages are configurable in config.json. Applied on "
            "Worksheet 7 (Episode Summary).",
            "body",
        ),
        ("", ""),
        ("KEY TOTALS — ANALYTICAL VIEW", "section"),
        (f"Annual Medicaid actual spend:           ${totals['actual']:>16,.0f}", "body"),
        (
            f"Annual measured outlier excess:         ${totals['outlier_excess']:>16,.0f}  "
            f"({totals['outlier_excess']/totals['actual']*100:.1f}% of spend)",
            "body",
        ),
        (
            f"Annual addressable (after classification): ${totals['addressable']:>13,.0f}  "
            f"({totals['addressable']/totals['outlier_excess']*100:.1f}% of outlier)",
            "body",
        ),
        (f"Facility-episodes in analytical view:   {totals['n_filtered']:>16,}", "body"),
        (
            f"Facility-episodes at O:E > {config['oe_outlier_threshold']}: "
            f"{totals['n_outliers']:>16,}  "
            f"({totals['n_outliers']/totals['n_filtered']*100:.1f}% of analytical set)",
            "body",
        ),
        (f"Unique facilities in scope:             {totals['n_facilities']:>16,}", "body"),
        (f"Episodes:                               {totals['n_episodes']:>16,}", "body"),
    ]

    styles = {
        "title": (Font(bold=True, size=18, color=NAVY), 40),
        "subtitle": (Font(italic=True, size=12, color="4A5A6A"), 18),
        "meta": (Font(size=9, color="8A9AAA"), 14),
        "section": (Font(bold=True, size=11, color="FFFFFF"), 22),
        "body": (Font(size=10), 16),
        "": (Font(size=10), 10),
    }

    section_fill = PatternFill(
        start_color=NAVY_DK, end_color=NAVY_DK, fill_type="solid"
    )

    row = 1
    for text, key in content:
        cell = ws.cell(row=row, column=1, value=text)
        font, height = styles.get(key, styles["body"])
        cell.font = font
        if key == "section":
            cell.fill = section_fill
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = height
        row += 1

    ws.column_dimensions["A"].width = 110
    for r in (1, 2, 3):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    for r_idx, (_, key) in enumerate(content, start=1):
        if key == "section":
            ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=6)


def build_all_data(wb, sorted_aug, new_header):
    ws = wb.create_sheet("All Data")
    ws.append(new_header)
    style_header_row(ws, 1, len(new_header))
    out_fill = PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid")
    lo_fill = PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type="solid")
    for r in sorted_aug:
        ws.append(r)
        is_outlier = r[-2] == "Yes"
        is_lo = "low_vol" in r[-1] or "very_low_vol" in r[-1] or "zero_vol" in r[-1]
        if is_outlier and not is_lo:
            for c in range(1, len(new_header) + 1):
                ws.cell(row=ws.max_row, column=c).fill = out_fill
        elif is_lo:
            for c in range(1, len(new_header) + 1):
                ws.cell(row=ws.max_row, column=c).fill = lo_fill
    ws.freeze_panes = "A2"
    autosize_columns(ws, max_width=30)
    ws.column_dimensions["B"].width = 50


def build_analytical_view(wb, filtered_aug, new_header, config):
    ws = wb.create_sheet("Analytical View")
    text = (
        f"FILTER: Medicaid Volume >= {config['min_volume_filter']}  ·  "
        f"{len(filtered_aug)} facility-episodes  ·  "
        "Suppresses small-n statistical noise"
    )
    filter_banner(ws, text, YELLOW_BG, "7A5200", len(new_header))
    ws.append(new_header)
    style_header_row(ws, 2, len(new_header))
    out_fill = PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid")
    for r in filtered_aug:
        ws.append(r)
        if r[-2] == "Yes":
            for c in range(1, len(new_header) + 1):
                ws.cell(row=ws.max_row, column=c).fill = out_fill
    ws.freeze_panes = "A3"
    autosize_columns(ws, max_width=30)
    ws.column_dimensions["B"].width = 50


def build_outliers_only(wb, outliers, new_header, total_excess, config):
    ws = wb.create_sheet("Outliers Only")
    text = (
        f"FILTER: vol >= {config['min_volume_filter']}  AND  "
        f"O:E > {config['oe_outlier_threshold']}  ·  {len(outliers)} facility-episodes "
        f"·  ${total_excess:,.0f} total outlier excess"
    )
    filter_banner(ws, text, RED_BG, "7F0000", len(new_header))
    ws.append(new_header)
    style_header_row(ws, 2, len(new_header))
    for r in outliers:
        ws.append(r)
    ws.freeze_panes = "A3"
    autosize_columns(ws, max_width=30)
    ws.column_dimensions["B"].width = 50


def build_systemic_outliers(wb, filtered_aug, col_idx, config):
    ws = wb.create_sheet("Systemic Outliers")
    text = (
        f"FILTER: vol >= {config['min_volume_filter']}  ·  "
        f"Facilities that are outlier (O:E > {config['oe_outlier_threshold']}) "
        f"in {config['systemic_min_episodes']}+ distinct episodes"
    )
    header = [
        "Facility",
        "Hospital System",
        "County",
        "DOI",
        "RAE Region",
        "Outlier Episodes (count)",
        "Total Episodes (count)",
        "Outlier Persistence %",
        "Outlier Excess $",
        "Outlier in (episode list)",
    ]
    filter_banner(ws, text, LIGHT_BG, NAVY, len(header))
    ws.append(header)
    style_header_row(ws, 2, len(header))

    fac_idx = col_idx["Facility Name"]
    sys_idx = col_idx["Facility Hospital System"]
    ep_idx = col_idx["Episode Name"]
    rae_idx = col_idx["Facility RAE Region"]
    doi_idx = col_idx["Facility DOI Category"]
    actual_idx = col_idx["Total Actual Cost"]
    oe_idx = col_idx["Facility O:E Cost Ratio"]
    county_idx = col_idx.get("Facility County")

    fac_eps = defaultdict(set)
    fac_out_eps = defaultdict(set)
    fac_excess = defaultdict(float)
    fac_meta = {}
    for r in filtered_aug:
        fac = r[fac_idx].strip()
        ep = r[ep_idx].strip()
        oe = parse_number(r[oe_idx])
        fac_eps[fac].add(ep)
        if fac not in fac_meta:
            fac_meta[fac] = {
                "system": r[sys_idx],
                "county": r[county_idx] if county_idx is not None else "",
                "doi": r[doi_idx],
                "rae": r[rae_idx],
            }
        if oe > config["oe_outlier_threshold"]:
            fac_out_eps[fac].add(ep)
            fac_excess[fac] += parse_number(r[actual_idx]) * (1 - 1 / oe)

    systemic = []
    for fac, eps in fac_out_eps.items():
        if len(eps) >= config["systemic_min_episodes"]:
            m = fac_meta[fac]
            systemic.append(
                (
                    fac,
                    m["system"],
                    m["county"],
                    m["doi"],
                    m["rae"],
                    len(eps),
                    len(fac_eps[fac]),
                    f"{len(eps)/len(fac_eps[fac])*100:.0f}%",
                    fac_excess[fac],
                    ", ".join(sorted(eps)),
                )
            )
    systemic.sort(key=lambda x: (-x[5], -x[8]))

    for row in systemic:
        ws.append(row)
        # Highlight by persistence
        if row[5] >= 5:
            for c in range(1, len(header) + 1):
                ws.cell(row=ws.max_row, column=c).fill = PatternFill(
                    start_color=RED_BG, end_color=RED_BG, fill_type="solid"
                )
        elif row[5] >= 4:
            for c in range(1, len(header) + 1):
                ws.cell(row=ws.max_row, column=c).fill = PatternFill(
                    start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type="solid"
                )
        # Currency
        ws.cell(row=ws.max_row, column=9).number_format = '"$"#,##0'

    ws.freeze_panes = "A3"
    widths = [50, 22, 14, 14, 10, 18, 18, 18, 16, 70]
    for col_l, w in zip("ABCDEFGHIJ", widths):
        ws.column_dimensions[col_l].width = w


def build_top_n_concentration(wb, outliers, col_idx, total_excess, config):
    N = config["top_n_concentration"]
    ws = wb.create_sheet(f"Top-{N} Concentration")
    cum_at_N = sum(
        parse_number(r[-3]) for r in outliers[:N]
    )
    pct_at_N = cum_at_N / total_excess * 100 if total_excess else 0
    text = (
        f"FILTER: vol >= {config['min_volume_filter']}, "
        f"O:E > {config['oe_outlier_threshold']}  ·  Top {N} facility-episodes by outlier excess $  "
        f"·  Captures {pct_at_N:.1f}% of ${total_excess:,.0f}"
    )
    header = [
        "Rank",
        "Facility",
        "Hospital System",
        "Episode",
        "Episode Classification",
        "Medicaid Volume",
        "% High Risk",
        "Facility O:E",
        "Outlier Excess $",
        "Cumulative Excess $",
        "Cumulative %",
    ]
    filter_banner(ws, text, LIGHT_BG, NAVY, len(header))
    ws.append(header)
    style_header_row(ws, 2, len(header))

    fac_idx = col_idx["Facility Name"]
    sys_idx = col_idx["Facility Hospital System"]
    ep_idx = col_idx["Episode Name"]
    vol_idx = col_idx["Medicaid Volume"]
    oe_idx = col_idx["Facility O:E Cost Ratio"]

    cumulative = 0
    for i, r in enumerate(outliers[:N], start=1):
        excess_val = parse_number(r[-3])
        cumulative += excess_val
        # The augmented columns are at the END, in this order:
        # [classification, %HR, rank, excess, isOutlier, dqFlag]
        # Indexed from the end: -6, -5, -4, -3, -2, -1
        ws.append(
            [
                i,
                r[fac_idx],
                r[sys_idx],
                r[ep_idx],
                r[-6],  # classification
                int(parse_number(r[vol_idx])),
                r[-5],  # %HR
                parse_number(r[oe_idx]),
                excess_val,
                cumulative,
                f"{cumulative/total_excess*100:.1f}%" if total_excess else "0%",
            ]
        )
        if i <= 10:
            for c in range(1, len(header) + 1):
                ws.cell(row=ws.max_row, column=c).fill = PatternFill(
                    start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid"
                )
        elif i <= 20:
            for c in range(1, len(header) + 1):
                ws.cell(row=ws.max_row, column=c).fill = PatternFill(
                    start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid"
                )
        ws.cell(row=ws.max_row, column=9).number_format = '"$"#,##0'
        ws.cell(row=ws.max_row, column=10).number_format = '"$"#,##0'

    ws.freeze_panes = "A3"
    widths = [6, 50, 22, 50, 18, 14, 12, 12, 18, 18, 14]
    for col_l, w in zip("ABCDEFGHIJK", widths):
        ws.column_dimensions[col_l].width = w


def build_episode_summary(wb, filtered_aug, col_idx, config):
    ws = wb.create_sheet("Episode Summary")
    text = (
        f"FILTER: vol >= {config['min_volume_filter']}  ·  "
        "Episode-level rollup with classification and addressable excess"
    )
    header = [
        "Episode",
        "Classification",
        "Addressable %",
        "Facilities (n)",
        "Outlier Facilities (n)",
        "Outlier Rate %",
        "Medicaid Volume",
        "Annual Actual Spend",
        "Outlier Excess $",
        "Addressable Excess $",
    ]
    filter_banner(ws, text, LIGHT_BG, NAVY, len(header))
    ws.append(header)
    style_header_row(ws, 2, len(header))

    ep_idx = col_idx["Episode Name"]
    vol_idx = col_idx["Medicaid Volume"]
    actual_idx = col_idx["Total Actual Cost"]
    oe_idx = col_idx["Facility O:E Cost Ratio"]

    ep_data = defaultdict(
        lambda: {"actual": 0, "vol": 0, "outlier_excess": 0, "n": 0, "n_outlier": 0}
    )
    for r in filtered_aug:
        ep = r[ep_idx].strip()
        vol = parse_number(r[vol_idx])
        actual = parse_number(r[actual_idx])
        oe = parse_number(r[oe_idx])
        ep_data[ep]["actual"] += actual
        ep_data[ep]["vol"] += vol
        ep_data[ep]["n"] += 1
        if oe > config["oe_outlier_threshold"]:
            ep_data[ep]["outlier_excess"] += actual * (1 - 1 / oe)
            ep_data[ep]["n_outlier"] += 1

    addr_map = config["addressable_pct_by_classification"]
    tot_act = tot_vol = tot_n = tot_n_out = tot_oe = tot_addr = 0
    for ep, d in sorted(ep_data.items(), key=lambda x: -x[1]["outlier_excess"]):
        cls = config["episode_classifications"].get(
            ep, config["default_classification_for_unknown"]
        )
        addr_pct = addr_map.get(cls, 0.5)
        addr_exc = d["outlier_excess"] * addr_pct
        out_rate = d["n_outlier"] / d["n"] * 100 if d["n"] else 0
        ws.append(
            [
                ep,
                cls,
                f"{int(addr_pct*100)}%",
                d["n"],
                d["n_outlier"],
                f"{out_rate:.1f}%",
                int(d["vol"]),
                d["actual"],
                d["outlier_excess"],
                addr_exc,
            ]
        )
        tot_act += d["actual"]
        tot_vol += d["vol"]
        tot_n += d["n"]
        tot_n_out += d["n_outlier"]
        tot_oe += d["outlier_excess"]
        tot_addr += addr_exc

    # Totals row
    ws.append(
        [
            f"TOTAL — {len(ep_data)} episodes",
            "",
            "",
            tot_n,
            tot_n_out,
            f"{tot_n_out/tot_n*100:.1f}%" if tot_n else "0%",
            int(tot_vol),
            tot_act,
            tot_oe,
            tot_addr,
        ]
    )
    total_row = ws.max_row
    for c in range(1, len(header) + 1):
        ws.cell(row=total_row, column=c).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=total_row, column=c).fill = PatternFill(
            start_color=NAVY_DK, end_color=NAVY_DK, fill_type="solid"
        )

    # Currency
    for r in range(3, ws.max_row + 1):
        ws.cell(row=r, column=7).number_format = "#,##0"
        for col in (8, 9, 10):
            ws.cell(row=r, column=col).number_format = '"$"#,##0'

    ws.freeze_panes = "A3"
    ws.column_dimensions["A"].width = 60
    for col_l in "BCDEFGHIJ":
        ws.column_dimensions[col_l].width = 18

    return tot_oe  # used elsewhere


def build_by_rae(wb, filtered_aug, col_idx, total_excess, config):
    ws = wb.create_sheet("By RAE Region")
    text = (
        f"FILTER: vol >= {config['min_volume_filter']}  ·  "
        "Outlier excess by Regional Accountable Entity catchment"
    )
    header = [
        "RAE Region",
        "Descriptor",
        "Unique Facilities",
        "Medicaid Volume",
        "Outlier Facility-Episodes",
        "Outlier Excess $",
        "% of Total Outlier Excess",
    ]
    filter_banner(ws, text, LIGHT_BG, NAVY, len(header))
    ws.append(header)
    style_header_row(ws, 2, len(header))

    fac_idx = col_idx["Facility Name"]
    rae_idx = col_idx["Facility RAE Region"]
    vol_idx = col_idx["Medicaid Volume"]
    actual_idx = col_idx["Total Actual Cost"]
    oe_idx = col_idx["Facility O:E Cost Ratio"]

    rae_excess = defaultdict(float)
    rae_vol = defaultdict(float)
    rae_facs = defaultdict(set)
    rae_out = defaultdict(int)
    for r in filtered_aug:
        rae = r[rae_idx].strip()
        rae_vol[rae] += parse_number(r[vol_idx])
        rae_facs[rae].add(r[fac_idx].strip())
        oe = parse_number(r[oe_idx])
        if oe > config["oe_outlier_threshold"]:
            rae_excess[rae] += parse_number(r[actual_idx]) * (1 - 1 / oe)
            rae_out[rae] += 1

    descriptors = config["rae_descriptors"]
    for rae in sorted(rae_excess.keys()):
        pct = rae_excess[rae] / total_excess * 100 if total_excess else 0
        ws.append(
            [
                f"Region {rae}",
                descriptors.get(rae, ""),
                len(rae_facs[rae]),
                int(rae_vol[rae]),
                rae_out[rae],
                rae_excess[rae],
                f"{pct:.1f}%",
            ]
        )
        if pct >= 30:
            for c in range(1, len(header) + 1):
                ws.cell(row=ws.max_row, column=c).fill = PatternFill(
                    start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid"
                )

    for r in range(3, ws.max_row + 1):
        ws.cell(row=r, column=4).number_format = "#,##0"
        ws.cell(row=r, column=6).number_format = '"$"#,##0'

    ws.freeze_panes = "A3"
    widths = [12, 32, 16, 16, 22, 18, 22]
    for col_l, w in zip("ABCDEFG", widths):
        ws.column_dimensions[col_l].width = w


def build_by_doi(wb, filtered_aug, col_idx, total_excess, config):
    ws = wb.create_sheet("By DOI Urbanity")
    text = (
        f"FILTER: vol >= {config['min_volume_filter']}  ·  "
        "Outlier rate and excess by facility DOI urbanity category"
    )
    header = [
        "DOI Category",
        "Unique Facilities",
        "Medicaid Volume",
        "Facility-Episodes (total)",
        "Outlier Facility-Episodes",
        "Outlier Rate %",
        "Outlier Excess $",
        "% of Total Outlier Excess",
    ]
    filter_banner(ws, text, LIGHT_BG, NAVY, len(header))
    ws.append(header)
    style_header_row(ws, 2, len(header))

    fac_idx = col_idx["Facility Name"]
    doi_idx = col_idx["Facility DOI Category"]
    vol_idx = col_idx["Medicaid Volume"]
    actual_idx = col_idx["Total Actual Cost"]
    oe_idx = col_idx["Facility O:E Cost Ratio"]

    doi_excess = defaultdict(float)
    doi_vol = defaultdict(float)
    doi_facs = defaultdict(set)
    doi_out = defaultdict(int)
    doi_total = defaultdict(int)
    for r in filtered_aug:
        doi = r[doi_idx].strip()
        doi_vol[doi] += parse_number(r[vol_idx])
        doi_facs[doi].add(r[fac_idx].strip())
        doi_total[doi] += 1
        oe = parse_number(r[oe_idx])
        if oe > config["oe_outlier_threshold"]:
            doi_excess[doi] += parse_number(r[actual_idx]) * (1 - 1 / oe)
            doi_out[doi] += 1

    for doi in sorted(doi_excess.keys(), key=lambda d: -doi_excess[d]):
        out_rate = doi_out[doi] / doi_total[doi] * 100 if doi_total[doi] else 0
        pct = doi_excess[doi] / total_excess * 100 if total_excess else 0
        ws.append(
            [
                doi,
                len(doi_facs[doi]),
                int(doi_vol[doi]),
                doi_total[doi],
                doi_out[doi],
                f"{out_rate:.1f}%",
                doi_excess[doi],
                f"{pct:.1f}%",
            ]
        )
        if out_rate >= 50:
            for c in range(1, len(header) + 1):
                ws.cell(row=ws.max_row, column=c).fill = PatternFill(
                    start_color=RED_BG, end_color=RED_BG, fill_type="solid"
                )

    for r in range(3, ws.max_row + 1):
        ws.cell(row=r, column=3).number_format = "#,##0"
        ws.cell(row=r, column=7).number_format = '"$"#,##0'

    ws.freeze_panes = "A3"
    widths = [16, 16, 16, 22, 22, 16, 18, 22]
    for col_l, w in zip("ABCDEFGH", widths):
        ws.column_dimensions[col_l].width = w


def build_by_system(wb, filtered_aug, col_idx, total_excess, config):
    ws = wb.create_sheet("By Hospital System")
    text = (
        f"FILTER: vol >= {config['min_volume_filter']}  ·  "
        "Outlier excess by hospital system / independent affiliation"
    )
    header = [
        "Hospital System",
        "Unique Facilities",
        "Facility-Episodes",
        "Outlier Facility-Episodes",
        "Outlier Rate %",
        "Outlier Excess $",
        "% of Total",
    ]
    filter_banner(ws, text, LIGHT_BG, NAVY, len(header))
    ws.append(header)
    style_header_row(ws, 2, len(header))

    fac_idx = col_idx["Facility Name"]
    sys_idx = col_idx["Facility Hospital System"]
    actual_idx = col_idx["Total Actual Cost"]
    oe_idx = col_idx["Facility O:E Cost Ratio"]

    sys_excess = defaultdict(float)
    sys_facs = defaultdict(set)
    sys_out = defaultdict(int)
    sys_total = defaultdict(int)
    for r in filtered_aug:
        s = r[sys_idx].strip() or "(unclassified)"
        sys_facs[s].add(r[fac_idx].strip())
        sys_total[s] += 1
        oe = parse_number(r[oe_idx])
        if oe > config["oe_outlier_threshold"]:
            sys_excess[s] += parse_number(r[actual_idx]) * (1 - 1 / oe)
            sys_out[s] += 1

    cum_pct = 0
    for sname in sorted(sys_excess.keys(), key=lambda s: -sys_excess[s]):
        out_rate = sys_out[sname] / sys_total[sname] * 100 if sys_total[sname] else 0
        pct = sys_excess[sname] / total_excess * 100 if total_excess else 0
        cum_pct += pct
        ws.append(
            [
                sname,
                len(sys_facs[sname]),
                sys_total[sname],
                sys_out[sname],
                f"{out_rate:.1f}%",
                sys_excess[sname],
                f"{pct:.1f}%",
            ]
        )
        if cum_pct <= 90 and pct >= 5:
            for c in range(1, len(header) + 1):
                ws.cell(row=ws.max_row, column=c).fill = PatternFill(
                    start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid"
                )

    for r in range(3, ws.max_row + 1):
        ws.cell(row=r, column=6).number_format = '"$"#,##0'

    ws.freeze_panes = "A3"
    widths = [28, 18, 18, 22, 16, 18, 14]
    for col_l, w in zip("ABCDEFG", widths):
        ws.column_dimensions[col_l].width = w


# ────────────────────────────────────────────────────────────────────────────
# MAIN ORCHESTRATION
# ────────────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Generate the 10-worksheet CFO workbook from a Colorado APCD CSV.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("input_csv", help="Path to input CSV file")
    parser.add_argument(
        "--output",
        "-o",
        help="Output XLSX path (default: input filename with _workbook.xlsx)",
    )
    parser.add_argument(
        "--config",
        "-c",
        help="Optional path to config.json (default: looks for config.json beside script)",
    )
    parser.add_argument(
        "--min-volume",
        type=int,
        help="Minimum Medicaid Volume filter (override config)",
    )
    parser.add_argument(
        "--oe-threshold",
        type=float,
        help="Outlier O:E threshold (override config)",
    )
    args = parser.parse_args()

    # Resolve paths
    input_path = Path(args.input_csv)
    if not input_path.exists():
        print(f"ERROR: Input file not found: {input_path}", file=sys.stderr)
        sys.exit(2)

    output_path = (
        Path(args.output)
        if args.output
        else input_path.with_name(input_path.stem + "_workbook.xlsx")
    )

    config_path = (
        Path(args.config)
        if args.config
        else Path(__file__).parent / "config.json"
    )

    print(f"\nPACES CFO Workbook Generator")
    print(f"  Input:  {input_path}")
    print(f"  Output: {output_path}")

    config = load_config(config_path)

    # Apply CLI overrides
    if args.min_volume is not None:
        config["min_volume_filter"] = args.min_volume
    if args.oe_threshold is not None:
        config["oe_outlier_threshold"] = args.oe_threshold

    print(
        f"  Filter: vol >= {config['min_volume_filter']}, "
        f"O:E > {config['oe_outlier_threshold']}"
    )

    # Read + validate
    with open(input_path, encoding="utf-8-sig", newline="") as f:
        first_line = next(csv.reader(f))
    col_idx = validate_schema(first_line, REQUIRED_COLUMNS)

    header, data = read_input_csv(input_path, col_idx)
    print(f"  Loaded {len(data):,} non-blank data rows from input")

    warn_unclassified_episodes(data, col_idx, config)

    # Compute O:E ranks within episode
    oe_ranks = compute_oe_ranks_within_episode(data, col_idx)

    # Augment all rows with computed columns
    augmented = [
        augment_row(i, r, col_idx, config, oe_ranks) for i, r in enumerate(data)
    ]
    added_cols = [
        "Episode Classification",
        "% High Risk Volume",
        "O:E Rank (within Episode)",
        "Outlier Excess $",
        "Is Outlier (O:E>threshold)",
        "Data Quality Flag",
    ]
    new_header = header + added_cols

    # Sort augmented: by total episode spend desc, then within episode by O:E desc
    actual_idx = col_idx["Total Actual Cost"]
    ep_spend = defaultdict(float)
    for r in augmented:
        ep_spend[r[col_idx["Episode Name"]].strip()] += parse_number(r[actual_idx])
    sorted_aug = sorted(
        augmented,
        key=lambda r: (
            -ep_spend[r[col_idx["Episode Name"]].strip()],
            r[col_idx["Episode Name"]],
            -parse_number(r[col_idx["Facility O:E Cost Ratio"]]),
            -parse_number(r[col_idx["Medicaid Volume"]]),
        ),
    )

    # Apply analytical filter
    min_vol = config["min_volume_filter"]
    vol_idx = col_idx["Medicaid Volume"]
    filtered_aug = [r for r in sorted_aug if parse_number(r[vol_idx]) >= min_vol]
    print(f"  Analytical view (vol >= {min_vol}): {len(filtered_aug):,} rows")

    # Identify outliers — sort by computed excess descending
    outliers = [r for r in filtered_aug if r[-2] == "Yes"]
    outliers.sort(key=lambda r: -parse_number(r[-3]))
    total_excess = sum(parse_number(r[-3]) for r in outliers)
    print(
        f"  Outliers at O:E > {config['oe_outlier_threshold']}: "
        f"{len(outliers):,}  (${total_excess:,.0f} total excess)"
    )

    # Compute totals for README
    fac_idx = col_idx["Facility Name"]
    ep_idx = col_idx["Episode Name"]
    totals = {
        "actual": sum(parse_number(r[actual_idx]) for r in filtered_aug),
        "outlier_excess": total_excess,
        "addressable": sum(
            parse_number(r[-3])
            * config["addressable_pct_by_classification"].get(r[-6], 0.5)
            for r in outliers
        ),
        "n_filtered": len(filtered_aug),
        "n_outliers": len(outliers),
        "n_facilities": len({r[fac_idx].strip() for r in filtered_aug}),
        "n_episodes": len({r[ep_idx].strip() for r in filtered_aug}),
    }

    # ─── BUILD WORKBOOK ───
    wb = Workbook()
    print(f"\nBuilding worksheets...")
    build_readme(wb, totals, config)
    print(f"  1. README")
    build_all_data(wb, sorted_aug, new_header)
    print(f"  2. All Data ({len(sorted_aug):,} rows)")
    build_analytical_view(wb, filtered_aug, new_header, config)
    print(f"  3. Analytical View ({len(filtered_aug):,} rows)")
    build_outliers_only(wb, outliers, new_header, total_excess, config)
    print(f"  4. Outliers Only ({len(outliers):,} rows)")
    build_systemic_outliers(wb, filtered_aug, col_idx, config)
    print(f"  5. Systemic Outliers")
    build_top_n_concentration(wb, outliers, col_idx, total_excess, config)
    print(f"  6. Top-{config['top_n_concentration']} Concentration")
    build_episode_summary(wb, filtered_aug, col_idx, config)
    print(f"  7. Episode Summary")
    build_by_rae(wb, filtered_aug, col_idx, total_excess, config)
    print(f"  8. By RAE Region")
    build_by_doi(wb, filtered_aug, col_idx, total_excess, config)
    print(f"  9. By DOI Urbanity")
    build_by_system(wb, filtered_aug, col_idx, total_excess, config)
    print(f" 10. By Hospital System")

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"\nDone. Saved workbook to:\n  {output_path}\n")


if __name__ == "__main__":
    main()
